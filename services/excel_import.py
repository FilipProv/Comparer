"""
Excel import service.
Parses an uploaded .xlsx file and returns a list of dicts ready for DB insertion.
Expected columns (case-insensitive, Polish or English headers):
  produkt / product, dostawca / supplier, ilość / quantity,
  jednostka / unit, cena / price, waluta / currency,
  kategoria / category, ważna_do / valid_until, uwagi / notes
"""

import io
import re
from datetime import date
from typing import Any
import pandas as pd

# ── Canonical field metadata ────────────────────────────────────────────────

# All fields the system understands
OUR_FIELDS = [
    {"key": "product_name",  "label": "Produkt",            "required": True},
    {"key": "supplier",      "label": "Dostawca",           "required": True},
    {"key": "category",      "label": "Kategoria",          "required": True},
    {"key": "quantity",      "label": "Ilość w ofercie",    "required": True},
    {"key": "unit",          "label": "Jednostka",          "required": True},
    {"key": "price_original","label": "Cena",               "required": True},
    {"key": "currency",      "label": "Waluta",             "required": True},
    {"key": "moq",           "label": "MOQ",                "required": False},
    {"key": "valid_until",   "label": "Ważna do",           "required": False},
    {"key": "spec_label",    "label": "Spec / forma",       "required": False},
    {"key": "notes",         "label": "Uwagi",              "required": False},
    {"key": "ref_number",    "label": "Nr oferty / ref.",   "required": False},
    {"key": "tier2_qty",     "label": "Próg 2 — ilość",     "required": False},
    {"key": "tier2_price",   "label": "Próg 2 — cena",      "required": False},
    {"key": "tier3_qty",          "label": "Próg 3 — ilość",          "required": False},
    {"key": "tier3_price",        "label": "Próg 3 — cena",            "required": False},
    {"key": "incoterm",           "label": "Incoterm (DDP/EXW/FOB…)", "required": False},
    {"key": "logistics_cost_pln", "label": "Koszt logistyki PLN/jedn.","required": False},
    {"key": "_skip",              "label": "— Pomiń tę kolumnę —",     "required": False},
]

# Keywords that hint at each field (lowercase, partial match)
_FIELD_KEYWORDS: dict[str, list[str]] = {
    "product_name":   ["produkt", "product", "substance", "substancja", "nazwa", "name", "ingredient", "skladnik", "material", "item"],
    "supplier":       ["dostawca", "supplier", "vendor", "producer", "firma", "company", "lieferant"],
    "category":       ["kategoria", "category", "typ", "type", "group", "grupa"],
    "quantity":       ["ilosc", "quantity", "qty", "amount", "menge", "volume", "pack size", "size"],
    "unit":           ["jednostka", "unit", "uom", "einheit", "jm"],
    "price_original": ["cena", "price", "preis", "prix", "koszt", "cost", "rate", "value", "kwota"],
    "currency":       ["waluta", "currency", "curr", "waehrung", "devise"],
    "moq":            ["moq", "minimum", "min order", "min.order", "min qty", "minimalna", "mindestbestellung"],
    "valid_until":    ["wazna", "valid", "expir", "validity", "bis", "do dnia", "date"],
    "spec_label":     ["spec", "forma", "form", "grade", "standard", "quality", "jakos", "purity", "czystosc", "stezenie", "koncentrat"],
    "notes":          ["uwagi", "notes", "remarks", "comment", "opis", "description", "info", "bemerk"],
    "ref_number":     ["ref", "numer", "number", "oferta", "quote", "nr ", "id", "kod"],
    "tier2_qty":      ["prog 2", "tier 2", "tier2", "qty2", "ilosc2", "quantity 2", "2nd"],
    "tier2_price":    ["cena2", "price2", "tier2 price", "prog 2 cena"],
    "tier3_qty":          ["prog 3", "tier 3", "tier3", "qty3", "ilosc3", "quantity 3", "3rd"],
    "tier3_price":        ["cena3", "price3", "tier3 price", "prog 3 cena"],
    "incoterm":           ["incoterm", "inco", "delivery term", "warunki dostawy", "terms", "delivery", "lieferbedingung", "ddp", "exw", "fob", "cif", "dap"],
    "logistics_cost_pln": ["logistyka", "logistics", "transport", "freight", "fracht", "shipping", "dostawa koszt", "koszt dostawy", "extra cost"],
}

# Mapping of accepted column aliases → canonical name (for standard import)
COLUMN_ALIASES: dict[str, str] = {
    "produkt": "product_name",
    "product": "product_name",
    "nazwa": "product_name",
    "substance": "product_name",
    "substancja": "product_name",
    "ingredient": "product_name",
    "dostawca": "supplier",
    "supplier": "supplier",
    "vendor": "supplier",
    "ilość": "quantity",
    "ilosc": "quantity",
    "quantity": "quantity",
    "qty": "quantity",
    "jednostka": "unit",
    "unit": "unit",
    "uom": "unit",
    "cena": "price_original",
    "price": "price_original",
    "cena_jedn": "price_original",
    "koszt": "price_original",
    "waluta": "currency",
    "currency": "currency",
    "curr": "currency",
    "kategoria": "category",
    "category": "category",
    "ważna_do": "valid_until",
    "wazna_do": "valid_until",
    "valid_until": "valid_until",
    "valid": "valid_until",
    "expiry": "valid_until",
    "data_waznosci": "valid_until",
    "uwagi": "notes",
    "notes": "notes",
    "remarks": "notes",
    "comments": "notes",
    "moq": "moq",
    "minimum": "moq",
    "min_order": "moq",
    "spec": "spec_label",
    "spec_label": "spec_label",
    "grade": "spec_label",
    "forma": "spec_label",
    "ref": "ref_number",
    "nr_oferty": "ref_number",
    "incoterm": "incoterm",
    "inco": "incoterm",
    "delivery_terms": "incoterm",
    "warunki_dostawy": "incoterm",
    "logistyka": "logistics_cost_pln",
    "transport": "logistics_cost_pln",
    "freight": "logistics_cost_pln",
    "koszt_dostawy": "logistics_cost_pln",
}

REQUIRED_COLUMNS = {"product_name", "supplier", "quantity", "unit", "price_original", "currency", "category"}

VALID_CATEGORIES = {"substancja_czynna", "opakowanie", "kapsula"}
VALID_CURRENCIES = {"PLN", "EUR", "USD"}


def _normalise_category(raw: str) -> str:
    mapping = {
        "substancja czynna": "substancja_czynna",
        "substancja_czynna": "substancja_czynna",
        "active substance": "substancja_czynna",
        "opakowanie": "opakowanie",
        "packaging": "opakowanie",
        "kapsula": "kapsula",
        "kapsułka": "kapsula",
        "capsule": "kapsula",
    }
    return mapping.get(str(raw).strip().lower(), str(raw).strip().lower())


def _parse_date(val: Any) -> date | None:
    if pd.isna(val) or val == "" or val is None:
        return None
    if isinstance(val, (date,)):
        return val
    try:
        return pd.to_datetime(str(val), dayfirst=True).date()
    except Exception:
        return None


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse Excel bytes.
    Returns (rows, errors) where rows are dicts ready for QuotationCreate,
    and errors is a list of human-readable problem descriptions.
    """
    errors: list[str] = []
    rows: list[dict] = []

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as exc:
        return [], [f"Nie można odczytać pliku Excel: {exc}"]

    # Normalise column names
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    rename_map = {col: COLUMN_ALIASES[col] for col in df.columns if col in COLUMN_ALIASES}
    df = df.rename(columns=rename_map)

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        return [], [f"Brakujące kolumny w pliku: {', '.join(missing)}"]

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # Excel row number (1-based + header)
        row_errors: list[str] = []

        product_name = str(row.get("product_name", "")).strip()
        supplier = str(row.get("supplier", "")).strip()
        unit = str(row.get("unit", "")).strip()
        currency = str(row.get("currency", "")).strip().upper()
        category = _normalise_category(row.get("category", ""))

        if not product_name:
            row_errors.append("brak nazwy produktu")
        if not supplier:
            row_errors.append("brak dostawcy")
        if category not in VALID_CATEGORIES:
            row_errors.append(f"nieznana kategoria '{category}'")
        if currency not in VALID_CURRENCIES:
            row_errors.append(f"nieznana waluta '{currency}'")

        try:
            quantity = float(str(row.get("quantity", "")).replace(",", "."))
            if quantity <= 0:
                raise ValueError
        except ValueError:
            quantity = 0.0
            row_errors.append("nieprawidłowa ilość")

        try:
            price_original = float(str(row.get("price_original", "")).replace(",", "."))
            if price_original <= 0:
                raise ValueError
        except ValueError:
            price_original = 0.0
            row_errors.append("nieprawidłowa cena")

        valid_until = _parse_date(row.get("valid_until"))
        notes = str(row.get("notes", "")).strip() or None

        if row_errors:
            errors.append(f"Wiersz {row_num}: {'; '.join(row_errors)}")
            continue

        # Parse optional fields
        try:
            moq = float(str(row.get("moq", "")).replace(",", ".")) if str(row.get("moq", "")).strip() else None
        except ValueError:
            moq = None

        try:
            logistic_raw = str(row.get("logistics_cost_pln", "")).strip()
            logistics_cost_pln = float(logistic_raw.replace(",", ".")) if logistic_raw else None
        except ValueError:
            logistics_cost_pln = None

        incoterm_raw = str(row.get("incoterm", "")).strip().upper()
        incoterm = incoterm_raw if incoterm_raw else None

        spec_label = str(row.get("spec_label", "")).strip() or None
        ref_number = str(row.get("ref_number", "")).strip() or None

        # Build notes — merge ref_number into notes if present
        note_parts = [p for p in [notes, f"Ref: {ref_number}" if ref_number else None] if p]
        combined_notes = "; ".join(note_parts) or None

        rows.append(
            {
                "product_name":      product_name,
                "supplier":          supplier,
                "quantity":          quantity,
                "unit":              unit,
                "price_original":    price_original,
                "currency":          currency,
                "category":          category,
                "valid_until":       valid_until,
                "notes":             combined_notes,
                "moq":               moq,
                "spec_label":        spec_label,
                "incoterm":          incoterm,
                "logistics_cost_pln": logistics_cost_pln,
            }
        )

    return rows, errors


# ── Column auto-detection ────────────────────────────────────────────────────

def _score_column(col_raw: str, field_key: str) -> int:
    """Return a 0-100 match score between a raw column header and a field key."""
    col = col_raw.lower().strip()
    col_nospace = re.sub(r"[\s_\-/]", "", col)
    keywords = _FIELD_KEYWORDS.get(field_key, [])
    best = 0
    for kw in keywords:
        kw_nospace = re.sub(r"[\s_\-/]", "", kw)
        if col_nospace == kw_nospace:
            best = max(best, 100)
        elif kw_nospace in col_nospace or col_nospace in kw_nospace:
            best = max(best, 80)
        elif kw in col:
            best = max(best, 70)
        elif any(part in col for part in kw.split()):
            best = max(best, 50)
    return best


def detect_column_mapping(file_bytes: bytes) -> dict:
    """
    Read column headers from an uploaded Excel file and propose a mapping
    to our canonical fields with a confidence score (0-100).

    Returns:
      {
        "columns": ["Col A", "Col B", ...],   # original headers
        "mapping": {
          "Col A": {"field": "product_name", "score": 95},
          "Col B": {"field": "price_original", "score": 80},
          ...
        },
        "sample_rows": [[...], [...], [...]]  # up to 3 sample data rows
      }
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, nrows=10)
    except Exception as exc:
        raise ValueError(f"Nie można odczytać pliku Excel: {exc}")

    original_columns = list(df.columns)
    field_keys = [f["key"] for f in OUR_FIELDS]

    # For each source column find best matching field
    mapping: dict[str, dict] = {}
    used_fields: set[str] = set()

    # Score matrix: source_col → {field_key → score}
    score_matrix: dict[str, dict[str, int]] = {}
    for col in original_columns:
        scores = {fk: _score_column(col, fk) for fk in field_keys if fk != "_skip"}
        score_matrix[col] = scores

    # Greedy assignment: highest score wins, each field assigned at most once
    # (except _skip which is unlimited)
    assignments: list[tuple[str, str, int]] = []
    for col in original_columns:
        ranked = sorted(score_matrix[col].items(), key=lambda x: -x[1])
        for fk, sc in ranked:
            if sc == 0:
                break
            if fk not in used_fields:
                assignments.append((col, fk, sc))
                used_fields.add(fk)
                break
        else:
            assignments.append((col, "_skip", 0))

    for col, fk, sc in assignments:
        mapping[col] = {"field": fk, "score": sc}

    # Fill unmatched columns with _skip
    for col in original_columns:
        if col not in mapping:
            mapping[col] = {"field": "_skip", "score": 0}

    # Sample rows (up to 3, exclude NaN)
    sample_rows = []
    for _, row in df.head(3).iterrows():
        sample_rows.append([str(v) if not pd.isna(v) else "" for v in row])

    return {
        "columns": original_columns,
        "mapping": mapping,
        "sample_rows": sample_rows,
        "our_fields": OUR_FIELDS,
    }


def apply_column_mapping(file_bytes: bytes, mapping: dict[str, str]) -> tuple[list[dict], list[str]]:
    """
    Apply a confirmed column mapping to an Excel file and return parsed rows.
    mapping: { "Original Col Name": "our_field_key" }
    Fields mapped to "_skip" are ignored.
    """
    errors: list[str] = []
    rows: list[dict] = []

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as exc:
        return [], [f"Nie można odczytać pliku Excel: {exc}"]

    # Build reverse: our_field → first matching source col
    field_to_col: dict[str, str] = {}
    for src_col, our_field in mapping.items():
        if our_field and our_field != "_skip" and our_field not in field_to_col:
            if src_col in df.columns:
                field_to_col[our_field] = src_col

    def get(row, field: str, default=""):
        col = field_to_col.get(field)
        if col is None:
            return default
        v = row.get(col, default)
        return "" if pd.isna(v) or v == "nan" else str(v).strip()

    required_fields = [f["key"] for f in OUR_FIELDS if f["required"]]
    missing_required = [f for f in required_fields if f not in field_to_col]
    if missing_required:
        labels = {f["key"]: f["label"] for f in OUR_FIELDS}
        errors.append(f"Brakujące wymagane pola: {', '.join(labels.get(f, f) for f in missing_required)}")
        return [], errors

    for idx, row in df.iterrows():
        row_num = int(idx) + 2
        row_errors: list[str] = []

        product_name = get(row, "product_name")
        supplier     = get(row, "supplier")
        unit         = get(row, "unit")
        currency     = get(row, "currency").upper()
        category_raw = get(row, "category")
        category     = _normalise_category(category_raw) if category_raw else "substancja_czynna"

        if not product_name: row_errors.append("brak nazwy produktu")
        if not supplier:     row_errors.append("brak dostawcy")
        if currency not in VALID_CURRENCIES:
            row_errors.append(f"nieznana waluta '{currency}'")

        try:
            quantity = float(get(row, "quantity").replace(",", "."))
            if quantity <= 0: raise ValueError
        except ValueError:
            quantity = 0.0
            row_errors.append("nieprawidłowa ilość")

        try:
            price_original = float(get(row, "price_original").replace(",", "."))
            if price_original <= 0: raise ValueError
        except ValueError:
            price_original = 0.0
            row_errors.append("nieprawidłowa cena")

        try:
            moq_raw = get(row, "moq")
            moq = float(moq_raw.replace(",", ".")) if moq_raw else None
        except ValueError:
            moq = None

        try:
            logistic_raw = get(row, "logistics_cost_pln")
            logistics_cost_pln = float(logistic_raw.replace(",", ".")) if logistic_raw else None
        except ValueError:
            logistics_cost_pln = None

        incoterm_raw = get(row, "incoterm")
        incoterm = incoterm_raw.upper().strip() if incoterm_raw else None
        # Normalise known incoterms
        _INCOTERM_MAP = {
            "DELIVERED DUTY PAID": "DDP", "EX WORKS": "EXW",
            "FREE ON BOARD": "FOB", "COST INSURANCE FREIGHT": "CIF",
            "DELIVERED AT PLACE": "DAP", "COST AND FREIGHT": "CFR",
        }
        if incoterm in _INCOTERM_MAP:
            incoterm = _INCOTERM_MAP[incoterm]

        valid_until = _parse_date(get(row, "valid_until") or None)
        notes_raw   = get(row, "notes") or None
        spec_label  = get(row, "spec_label") or None
        ref_number  = get(row, "ref_number") or None

        # Tier 2 / Tier 3 price breaks — store in notes if provided
        tier_notes = []
        for tier_n in (2, 3):
            t_qty_raw   = get(row, f"tier{tier_n}_qty")
            t_price_raw = get(row, f"tier{tier_n}_price")
            if t_qty_raw and t_price_raw:
                try:
                    t_qty   = float(t_qty_raw.replace(",", "."))
                    t_price = float(t_price_raw.replace(",", "."))
                    tier_notes.append(f"Próg {tier_n}: {t_qty} {unit} → {t_price} {currency}")
                except ValueError:
                    pass

        note_parts = [p for p in [notes_raw, f"Ref: {ref_number}" if ref_number else None] + tier_notes if p]
        combined_notes = "; ".join(note_parts) or None

        if row_errors:
            errors.append(f"Wiersz {row_num}: {'; '.join(row_errors)}")
            continue

        rows.append({
            "product_name":      product_name,
            "supplier":          supplier,
            "quantity":          quantity,
            "unit":              unit,
            "price_original":    price_original,
            "currency":          currency,
            "category":          category,
            "valid_until":       valid_until,
            "notes":             combined_notes,
            "moq":               moq,
            "spec_label":        spec_label,
            "incoterm":          incoterm,
            "logistics_cost_pln": logistics_cost_pln,
        })

    return rows, errors


def apply_column_mapping_catalog(file_bytes: bytes, mapping: dict[str, str]) -> tuple[list[dict], list[str]]:
    """Like apply_column_mapping but only extracts catalog-relevant fields (no price required)."""
    errors: list[str] = []
    rows: list[dict] = []

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as exc:
        return [], [f"Nie można odczytać pliku Excel: {exc}"]

    field_to_col: dict[str, str] = {}
    for src_col, our_field in mapping.items():
        if our_field and our_field != "_skip" and our_field not in field_to_col:
            if src_col in df.columns:
                field_to_col[our_field] = src_col

    def get(row, field: str, default=""):
        col = field_to_col.get(field)
        if col is None:
            return default
        v = row.get(col, default)
        return "" if pd.isna(v) or v == "nan" else str(v).strip()

    for idx, row in df.iterrows():
        product_name = get(row, "product_name")
        supplier     = get(row, "supplier")
        if not product_name:
            continue  # silently skip blank rows
        category_raw = get(row, "category")
        category     = _normalise_category(category_raw) if category_raw else "substancja_czynna"
        notes_parts  = [p for p in [get(row, "notes"), get(row, "spec_label")] if p]
        rows.append({
            "product_name":  product_name,
            "supplier":      supplier,
            "category":      category,
            "contact_email": get(row, "contact_email") or None,
            "notes":         "; ".join(notes_parts) or None,
        })

    return rows, errors


def build_template_excel() -> bytes:
    """Generate a full-featured template Excel file for users to fill in."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl

    columns = [
        ("produkt",          "Vitamin C Ascorbic Acid"),
        ("dostawca",         "Firma X GmbH"),
        ("kategoria",        "substancja_czynna"),
        ("ilosc_w_ofercie",  "25"),
        ("jednostka",        "kg"),
        ("cena",             "12.50"),
        ("waluta",           "EUR"),
        ("moq",              "5"),
        ("wazna_do",         "2026-12-31"),
        ("spec_forma",       "USP, 99% pure"),
        ("uwagi",            "lead time 4-6 tygodni"),
        ("nr_oferty_ref",    "QT-2026-0412"),
        ("prog2_ilosc",      "50"),
        ("prog2_cena",       "11.20"),
        ("prog3_ilosc",      "100"),
        ("prog3_cena",       "10.00"),
        ("incoterm",         "DDP"),
        ("koszt_logistyki_pln_jedn", ""),
    ]

    HEADER_NOTES = {
        "produkt":         "Pełna nazwa składnika (np. Vitamin C Ascorbic Acid)",
        "dostawca":        "Nazwa dostawcy / firmy",
        "kategoria":       "substancja_czynna | opakowanie | kapsula",
        "ilosc_w_ofercie": "Ilość przy której obowiązuje podana cena",
        "jednostka":       "kg | g | mg | szt | op | l | ml",
        "cena":            "Tylko liczba (bez waluty)",
        "waluta":          "PLN | EUR | USD",
        "moq":             "Minimalna ilość zamówienia (ta sama jednostka)",
        "wazna_do":        "Format: RRRR-MM-DD",
        "spec_forma":      "np. USP, 50%, pure, micronized",
        "uwagi":           "Dodatkowe informacje (lead time, warunki itp.)",
        "nr_oferty_ref":   "Numer referencyjny oferty od dostawcy",
        "prog2_ilosc":     "Od jakiej ilości drugi próg cenowy",
        "prog2_cena":      "Cena w drugim progu (ta sama waluta)",
        "prog3_ilosc":     "Od jakiej ilości trzeci próg cenowy",
        "prog3_cena":      "Cena w trzecim progu (ta sama waluta)",
        "incoterm":        "Warunki dostawy: DDP (cena z dostawą) | EXW | FOB | CIF | DAP",
        "koszt_logistyki_pln_jedn": "Jeśli nie DDP: szacowany koszt dostawy PLN na jednostkę (do porównania)",
    }

    REQUIRED_COLS = {"produkt", "dostawca", "kategoria", "ilosc_w_ofercie", "jednostka", "cena", "waluta"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wyceny"

    # ── Style definitions ──
    hdr_fill_req  = PatternFill("solid", fgColor="1a6b3c")   # dark green = required
    hdr_fill_opt  = PatternFill("solid", fgColor="2c3e50")   # dark blue  = optional
    hdr_fill_tier = PatternFill("solid", fgColor="6f42c1")   # purple     = price tiers
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    data_fill = PatternFill("solid", fgColor="F0FFF4")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    TIER_COLS = {"prog2_ilosc", "prog2_cena", "prog3_ilosc", "prog3_cena"}

    for c_idx, (col_key, example_val) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_key)
        if col_key in TIER_COLS:
            cell.fill = hdr_fill_tier
        elif col_key in REQUIRED_COLS:
            cell.fill = hdr_fill_req
        else:
            cell.fill = hdr_fill_opt
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        if col_key in HEADER_NOTES:
            cell.comment = None  # openpyxl comment needs openpyxl.comments.Comment
            # Use a second row as description instead (more compatible)

        # Example data row
        data_cell = ws.cell(row=2, column=c_idx, value=example_val)
        data_cell.fill = data_fill
        data_cell.border = border
        data_cell.alignment = Alignment(horizontal="left")

        # Description row
        desc_cell = ws.cell(row=3, column=c_idx, value=HEADER_NOTES.get(col_key, ""))
        desc_cell.font = Font(color="888888", italic=True, size=8)
        desc_cell.alignment = Alignment(wrap_text=True, horizontal="left")
        desc_cell.border = border

    # Row heights
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 42

    # Column widths
    col_widths = [28, 22, 20, 16, 12, 10, 10, 10, 14, 20, 30, 18, 14, 12, 14, 12, 18, 24]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Legend sheet ──
    leg = wb.create_sheet("Legenda")
    leg["A1"] = "Kolor nagłówka"
    leg["B1"] = "Znaczenie"
    leg["A1"].font = Font(bold=True)
    leg["B1"].font = Font(bold=True)
    leg_data = [
        ("Zielony",   "Pole wymagane — bez niego import nie zadziała"),
        ("Ciemny",    "Pole opcjonalne — bardzo zalecane"),
        ("Fioletowy", "Progi cenowe — wpisz gdy dostawca daje rabat przy większym zamówieniu"),
    ]
    for r, (color, desc) in enumerate(leg_data, start=2):
        leg.cell(r, 1, color)
        leg.cell(r, 2, desc)
    for col_cells in leg.columns:
        leg.column_dimensions[col_cells[0].column_letter].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Export mapped rows as filled template ────────────────────────────────────

def build_filled_template(rows: list[dict], supplier_override: str | None = None) -> bytes:
    """
    Take a list of parsed rows (from apply_column_mapping) and produce
    an Excel file in our standard template format with all data pre-filled.
    Users can then edit manually before importing via the universal import tab.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import openpyxl

    # Column definition: (our_field_key, display_header, width)
    COLS = [
        ("product_name",      "produkt",               28),
        ("supplier",          "dostawca",               22),
        ("category",          "kategoria",              20),
        ("quantity",          "ilosc_w_ofercie",        16),
        ("unit",              "jednostka",              12),
        ("price_original",    "cena",                   12),
        ("currency",          "waluta",                 10),
        ("moq",               "moq",                    10),
        ("valid_until",       "wazna_do",               14),
        ("incoterm",          "incoterm",               14),
        ("logistics_cost_pln","koszt_logistyki_pln_jedn", 24),
        ("spec_label",        "spec_forma",             20),
        ("notes",             "uwagi",                  35),
    ]

    REQUIRED = {"produkt", "dostawca", "kategoria", "ilosc_w_ofercie", "jednostka", "cena", "waluta"}

    hdr_fill_req = PatternFill("solid", fgColor="1a6b3c")
    hdr_fill_opt = PatternFill("solid", fgColor="2c3e50")
    hdr_font     = Font(color="FFFFFF", bold=True, size=10)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alternating row colours
    fill_even = PatternFill("solid", fgColor="F0FFF4")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wyceny"

    # Header row
    for c_idx, (_, header, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = hdr_fill_req if header in REQUIRED else hdr_fill_opt
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        fill = fill_even if r_idx % 2 == 0 else fill_odd
        for c_idx, (field_key, _, _w) in enumerate(COLS, start=1):
            value = row.get(field_key)
            # Apply supplier override if field is supplier and value is missing
            if field_key == "supplier" and not value and supplier_override:
                value = supplier_override
            # Format date
            if field_key == "valid_until" and value and hasattr(value, "isoformat"):
                value = value.isoformat()
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r_idx].height = 16

    # Info sheet explaining how to import
    info = wb.create_sheet("Jak importować")
    info_rows = [
        ("Krok 1:", "Sprawdź i popraw dane w arkuszu 'Wyceny' — możesz edytować każdą komórkę."),
        ("Krok 2:", "Kategoria musi być jedną z: substancja_czynna | opakowanie | kapsula"),
        ("Krok 3:", "Waluta: PLN | EUR | USD"),
        ("Krok 4:", "Zapisz plik i wgraj go w zakładce Import → 'Własne pliki / Uniwersalny import'."),
        ("Uwaga:",  "Kolumny zielone = wymagane. Kolumny ciemne = opcjonalne ale zalecane."),
    ]
    info["A1"] = "Instrukcja importu"
    info["A1"].font = Font(bold=True, size=12)
    for r, (label, text) in enumerate(info_rows, start=2):
        info.cell(r, 1, label).font = Font(bold=True)
        info.cell(r, 2, text)
    info.column_dimensions["A"].width = 12
    info.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
